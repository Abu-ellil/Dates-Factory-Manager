import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'src'))

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import license_manager
import sqlite3
from datetime import datetime
import json

class LicenseKeyGenGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Date Factory Manager - License Key Generator")
        self.root.geometry("900x700")
        self.root.resizable(False, False)
        
        # Set icon if available
        icon_path = os.path.join(os.path.dirname(__file__), '..', 'icon.ico')
        if os.path.exists(icon_path):
            self.root.iconbitmap(icon_path)
        
        # Initialize database
        self.init_database()
        
        # Color scheme
        self.bg_color = "#1a1a2e"
        self.secondary_bg = "#16213e"
        self.accent_color = "#0f3460"
        self.highlight_color = "#e94560"
        self.text_color = "#ffffff"
        self.success_color = "#00ff88"
        
        self.root.configure(bg=self.bg_color)
        
        self.setup_ui()
        self.load_keys_history()
        
    def init_database(self):
        """Initialize SQLite database for storing generated keys"""
        db_path = os.path.join(os.path.dirname(__file__), '..', 'license_keys.db')
        self.conn = sqlite3.connect(db_path)
        self.cursor = self.conn.cursor()
        
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS generated_keys (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                machine_id TEXT NOT NULL,
                client_name TEXT NOT NULL,
                expiration_date TEXT,
                license_key TEXT NOT NULL,
                generated_at TEXT NOT NULL,
                notes TEXT
            )
        ''')
        self.conn.commit()
        
    def setup_ui(self):
        """Setup the user interface"""
        # Header
        header_frame = tk.Frame(self.root, bg=self.highlight_color, height=80)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(
            header_frame,
            text="🔑 License Key Generator",
            font=("Segoe UI", 24, "bold"),
            bg=self.highlight_color,
            fg=self.text_color
        )
        title_label.pack(pady=20)
        
        # Main container
        main_container = tk.Frame(self.root, bg=self.bg_color)
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Left panel - Key Generation
        left_panel = tk.Frame(main_container, bg=self.secondary_bg, relief=tk.RAISED, bd=2)
        left_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        self.setup_generation_panel(left_panel)
        
        # Right panel - History
        right_panel = tk.Frame(main_container, bg=self.secondary_bg, relief=tk.RAISED, bd=2)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        self.setup_history_panel(right_panel)
        
    def setup_generation_panel(self, parent):
        """Setup the key generation panel"""
        # Title
        title = tk.Label(
            parent,
            text="Generate New License",
            font=("Segoe UI", 16, "bold"),
            bg=self.secondary_bg,
            fg=self.text_color
        )
        title.pack(pady=15)
        
        # Form frame
        form_frame = tk.Frame(parent, bg=self.secondary_bg)
        form_frame.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)
        
        # Machine ID
        self.create_label(form_frame, "Machine ID:", 0)
        self.machine_id_entry = self.create_entry(form_frame, 1, "XXXX-XXXX-XXXX-XXXX")
        
        # Client Name
        self.create_label(form_frame, "Client Name:", 2)
        self.client_name_entry = self.create_entry(form_frame, 3, "Enter customer name")
        
        # Expiration Date
        self.create_label(form_frame, "Expiration Date:", 4)
        exp_frame = tk.Frame(form_frame, bg=self.secondary_bg)
        exp_frame.grid(row=5, column=0, pady=5, sticky="ew")
        
        self.exp_date_entry = self.create_entry(exp_frame, 0, "YYYY-MM-DD (optional)")
        self.exp_date_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        lifetime_btn = tk.Button(
            exp_frame,
            text="Lifetime",
            command=self.set_lifetime,
            bg=self.accent_color,
            fg=self.text_color,
            font=("Segoe UI", 9),
            relief=tk.FLAT,
            cursor="hand2"
        )
        lifetime_btn.pack(side=tk.RIGHT, padx=(5, 0))
        
        # Notes
        self.create_label(form_frame, "Notes (Optional):", 6)
        self.notes_text = tk.Text(
            form_frame,
            height=4,
            font=("Segoe UI", 10),
            bg=self.accent_color,
            fg=self.text_color,
            insertbackground=self.text_color,
            relief=tk.FLAT,
            padx=10,
            pady=8
        )
        self.notes_text.grid(row=7, column=0, pady=5, sticky="ew")
        
        # Generate Button
        generate_btn = tk.Button(
            form_frame,
            text="🔑 Generate License Key",
            command=self.generate_key,
            bg=self.highlight_color,
            fg=self.text_color,
            font=("Segoe UI", 12, "bold"),
            relief=tk.FLAT,
            cursor="hand2",
            height=2
        )
        generate_btn.grid(row=8, column=0, pady=20, sticky="ew")
        
        # Generated Key Display
        self.create_label(form_frame, "Generated License Key:", 9)
        
        key_display_frame = tk.Frame(form_frame, bg=self.accent_color, relief=tk.SUNKEN, bd=2)
        key_display_frame.grid(row=10, column=0, pady=5, sticky="ew")
        
        self.key_display = tk.Text(
            key_display_frame,
            height=3,
            font=("Courier New", 10, "bold"),
            bg=self.accent_color,
            fg=self.success_color,
            relief=tk.FLAT,
            padx=10,
            pady=8,
            wrap=tk.WORD
        )
        self.key_display.pack(fill=tk.BOTH, expand=True)
        self.key_display.config(state=tk.DISABLED)
        
        # Copy Button
        copy_btn = tk.Button(
            form_frame,
            text="📋 Copy to Clipboard",
            command=self.copy_to_clipboard,
            bg=self.accent_color,
            fg=self.text_color,
            font=("Segoe UI", 10),
            relief=tk.FLAT,
            cursor="hand2"
        )
        copy_btn.grid(row=11, column=0, pady=5, sticky="ew")
        
        form_frame.columnconfigure(0, weight=1)
        
    def setup_history_panel(self, parent):
        """Setup the history panel"""
        # Title
        title = tk.Label(
            parent,
            text="Generated Keys History",
            font=("Segoe UI", 16, "bold"),
            bg=self.secondary_bg,
            fg=self.text_color
        )
        title.pack(pady=15)
        
        # Search frame
        search_frame = tk.Frame(parent, bg=self.secondary_bg)
        search_frame.pack(padx=20, pady=5, fill=tk.X)
        
        search_label = tk.Label(
            search_frame,
            text="Search:",
            font=("Segoe UI", 10),
            bg=self.secondary_bg,
            fg=self.text_color
        )
        search_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.search_entry = tk.Entry(
            search_frame,
            font=("Segoe UI", 10),
            bg=self.accent_color,
            fg=self.text_color,
            insertbackground=self.text_color,
            relief=tk.FLAT
        )
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.search_entry.bind('<KeyRelease>', lambda e: self.search_keys())
        
        # Treeview
        tree_frame = tk.Frame(parent, bg=self.secondary_bg)
        tree_frame.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        # Configure style
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                       background=self.accent_color,
                       foreground=self.text_color,
                       fieldbackground=self.accent_color,
                       borderwidth=0)
        style.configure("Treeview.Heading",
                       background=self.highlight_color,
                       foreground=self.text_color,
                       borderwidth=0)
        style.map('Treeview', background=[('selected', self.highlight_color)])
        
        self.tree = ttk.Treeview(
            tree_frame,
            columns=("ID", "Client", "Machine ID", "Expiration", "Generated"),
            show="headings",
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set
        )
        
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        
        # Define columns
        self.tree.heading("ID", text="ID")
        self.tree.heading("Client", text="Client Name")
        self.tree.heading("Machine ID", text="Machine ID")
        self.tree.heading("Expiration", text="Expiration")
        self.tree.heading("Generated", text="Generated At")
        
        self.tree.column("ID", width=40, anchor="center")
        self.tree.column("Client", width=120)
        self.tree.column("Machine ID", width=140)
        self.tree.column("Expiration", width=90, anchor="center")
        self.tree.column("Generated", width=130, anchor="center")
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)
        
        # Bind double-click to view details
        self.tree.bind('<Double-1>', self.view_key_details)
        
        # Buttons frame
        btn_frame = tk.Frame(parent, bg=self.secondary_bg)
        btn_frame.pack(padx=20, pady=10, fill=tk.X)
        
        view_btn = tk.Button(
            btn_frame,
            text="👁 View Details",
            command=self.view_key_details,
            bg=self.accent_color,
            fg=self.text_color,
            font=("Segoe UI", 9),
            relief=tk.FLAT,
            cursor="hand2"
        )
        view_btn.pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        
        export_btn = tk.Button(
            btn_frame,
            text="📊 Export to Excel",
            command=self.export_to_excel,
            bg=self.accent_color,
            fg=self.text_color,
            font=("Segoe UI", 9),
            relief=tk.FLAT,
            cursor="hand2"
        )
        export_btn.pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        
        delete_btn = tk.Button(
            btn_frame,
            text="🗑 Delete",
            command=self.delete_key,
            bg=self.highlight_color,
            fg=self.text_color,
            font=("Segoe UI", 9),
            relief=tk.FLAT,
            cursor="hand2"
        )
        delete_btn.pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        
    def create_label(self, parent, text, row):
        """Create a styled label"""
        label = tk.Label(
            parent,
            text=text,
            font=("Segoe UI", 11),
            bg=self.secondary_bg,
            fg=self.text_color,
            anchor="w"
        )
        label.grid(row=row, column=0, pady=(10, 2), sticky="w")
        return label
        
    def create_entry(self, parent, row, placeholder=""):
        """Create a styled entry"""
        entry = tk.Entry(
            parent,
            font=("Segoe UI", 11),
            bg=self.accent_color,
            fg=self.text_color,
            insertbackground=self.text_color,
            relief=tk.FLAT
        )
        if isinstance(parent, tk.Frame) and hasattr(parent, 'grid'):
            entry.grid(row=row, column=0, pady=5, ipady=8, sticky="ew")
        return entry
        
    def set_lifetime(self):
        """Set expiration to lifetime"""
        self.exp_date_entry.delete(0, tk.END)
        self.exp_date_entry.insert(0, "LIFETIME")
        
    def generate_key(self):
        """Generate a new license key"""
        machine_id = self.machine_id_entry.get().strip().upper()
        client_name = self.client_name_entry.get().strip()
        exp_date = self.exp_date_entry.get().strip()
        notes = self.notes_text.get("1.0", tk.END).strip()
        
        # Validation
        if not machine_id:
            messagebox.showerror("Error", "Machine ID is required!")
            return
            
        if not client_name:
            messagebox.showerror("Error", "Client Name is required!")
            return
            
        # Handle expiration
        if exp_date.upper() == "LIFETIME" or not exp_date:
            exp_date = None
        else:
            # Validate date format
            try:
                datetime.strptime(exp_date, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("Error", "Invalid date format! Use YYYY-MM-DD")
                return
        
        try:
            # Generate license key
            license_key = license_manager.generate_license_key(machine_id, client_name, exp_date)
            
            # Display the key
            self.key_display.config(state=tk.NORMAL)
            self.key_display.delete("1.0", tk.END)
            self.key_display.insert("1.0", license_key)
            self.key_display.config(state=tk.DISABLED)
            
            # Save to database
            generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.cursor.execute('''
                INSERT INTO generated_keys (machine_id, client_name, expiration_date, license_key, generated_at, notes)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (machine_id, client_name, exp_date or "LIFETIME", license_key, generated_at, notes))
            self.conn.commit()
            
            # Refresh history
            self.load_keys_history()
            
            messagebox.showinfo("Success", f"License key generated successfully for {client_name}!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate key: {str(e)}")
            
    def copy_to_clipboard(self):
        """Copy the generated key to clipboard"""
        key = self.key_display.get("1.0", tk.END).strip()
        if key:
            self.root.clipboard_clear()
            self.root.clipboard_append(key)
            messagebox.showinfo("Copied", "License key copied to clipboard!")
        else:
            messagebox.showwarning("Warning", "No key to copy!")
            
    def load_keys_history(self):
        """Load all generated keys from database"""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # Load from database
        self.cursor.execute('''
            SELECT id, client_name, machine_id, expiration_date, generated_at
            FROM generated_keys
            ORDER BY id DESC
        ''')
        
        for row in self.cursor.fetchall():
            self.tree.insert("", tk.END, values=row)
            
    def search_keys(self):
        """Search keys by client name or machine ID"""
        search_term = self.search_entry.get().strip().lower()
        
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        if not search_term:
            self.load_keys_history()
            return
            
        # Search in database
        self.cursor.execute('''
            SELECT id, client_name, machine_id, expiration_date, generated_at
            FROM generated_keys
            WHERE LOWER(client_name) LIKE ? OR LOWER(machine_id) LIKE ?
            ORDER BY id DESC
        ''', (f'%{search_term}%', f'%{search_term}%'))
        
        for row in self.cursor.fetchall():
            self.tree.insert("", tk.END, values=row)
            
    def view_key_details(self, event=None):
        """View full details of a selected key"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a key to view!")
            return
            
        item = self.tree.item(selection[0])
        key_id = item['values'][0]
        
        # Get full details from database
        self.cursor.execute('''
            SELECT * FROM generated_keys WHERE id = ?
        ''', (key_id,))
        
        row = self.cursor.fetchone()
        if row:
            details = f"""
License Key Details
{'='*60}

ID: {row[0]}
Client Name: {row[2]}
Machine ID: {row[1]}
Expiration: {row[3]}
Generated At: {row[5]}

License Key:
{row[4]}

Notes:
{row[6] or 'No notes'}
            """
            
            # Create details window
            details_window = tk.Toplevel(self.root)
            details_window.title("License Key Details")
            details_window.geometry("600x500")
            details_window.configure(bg=self.bg_color)
            
            text_widget = tk.Text(
                details_window,
                font=("Courier New", 10),
                bg=self.accent_color,
                fg=self.text_color,
                padx=20,
                pady=20,
                wrap=tk.WORD
            )
            text_widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            text_widget.insert("1.0", details)
            text_widget.config(state=tk.DISABLED)
            
            # Copy button
            copy_btn = tk.Button(
                details_window,
                text="📋 Copy License Key",
                command=lambda: self.copy_key_from_details(row[4]),
                bg=self.highlight_color,
                fg=self.text_color,
                font=("Segoe UI", 11, "bold"),
                relief=tk.FLAT,
                cursor="hand2"
            )
            copy_btn.pack(pady=10)
            
    def copy_key_from_details(self, key):
        """Copy key from details window"""
        self.root.clipboard_clear()
        self.root.clipboard_append(key)
        messagebox.showinfo("Copied", "License key copied to clipboard!")
        
    def delete_key(self):
        """Delete a selected key"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a key to delete!")
            return
            
        item = self.tree.item(selection[0])
        key_id = item['values'][0]
        client_name = item['values'][1]
        
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete the key for '{client_name}'?"):
            self.cursor.execute('DELETE FROM generated_keys WHERE id = ?', (key_id,))
            self.conn.commit()
            self.load_keys_history()
            messagebox.showinfo("Deleted", "License key deleted successfully!")
            
    def export_to_excel(self):
        """Export all keys to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Ask for save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"license_keys_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not file_path:
                return
                
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "License Keys"
            
            # Headers
            headers = ["ID", "Client Name", "Machine ID", "Expiration Date", "License Key", "Generated At", "Notes"]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="E94560", end_color="E94560", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Get all data
            self.cursor.execute('SELECT * FROM generated_keys ORDER BY id DESC')
            
            for row in self.cursor.fetchall():
                ws.append(row)
                
            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 50
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 30
            
            # Save
            wb.save(file_path)
            messagebox.showinfo("Success", f"Keys exported successfully to:\n{file_path}")
            
        except ImportError:
            messagebox.showerror("Error", "openpyxl library is required for Excel export!\nInstall it using: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {str(e)}")
            
    def __del__(self):
        """Close database connection"""
        if hasattr(self, 'conn'):
            self.conn.close()

def main():
    root = tk.Tk()
    app = LicenseKeyGenGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
