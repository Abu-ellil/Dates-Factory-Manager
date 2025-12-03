"""
Test script to verify that the Excel import file locking issue has been fixed.
This script tests both the restore_data and bulk_import functionality to ensure
files are properly closed after processing.
"""
import os
import sys
import tempfile
from pathlib import Path

# Add src directory to Python path so imports work
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

def test_restore_data_file_handling():
    """Test that restore_data properly closes files after processing"""
    print("Testing restore_data file handling...")
    
    # Import here to avoid issues if modules don't exist
    try:
        from restore_data import restore_from_excel
        print("✓ Successfully imported restore_data module")
    except ImportError as e:
        print(f"✗ Failed to import restore_data module: {e}")
        return False
    
    # Create a temporary Excel file to test with
    import openpyxl
    test_workbook = openpyxl.Workbook()
    test_workbook.create_sheet('العملاء (Customers)')
    test_workbook.create_sheet('الميزان (Weighbridge)')
    test_workbook.create_sheet('الصناديق (Crates)')
    test_workbook.create_sheet('المالية (Finance)')
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        test_workbook.save(tmp.name)
        test_file_path = tmp.name
    
    try:
        # Test normal operation
        result = restore_from_excel(test_file_path, 'merge')
        print(f"✓ restore_from_excel completed with result: {result}")
        
        # Test that we can now delete the file (meaning it's not locked)
        os.remove(test_file_path)
        print("✓ Successfully deleted test file after restore operation (file was not locked)")
        return True
        
    except Exception as e:
        print(f"✗ Error during restore test: {e}")
        # Try to manually delete the file for cleanup
        try:
            os.remove(test_file_path)
        except:
            pass
        return False


def test_bulk_import_file_handling():
    """Test that bulk_import properly closes files after processing"""
    print("\nTesting bulk_import file handling...")
    
    # Import here to avoid issues if modules don't exist
    try:
        from bulk_import import import_customers_from_excel
        print("✓ Successfully imported bulk_import module")
    except ImportError as e:
        print(f"✗ Failed to import bulk_import module: {e}")
        return False
    
    # Create a temporary Excel file to test with
    import openpyxl
    test_workbook = openpyxl.Workbook()
    sheet = test_workbook.active
    sheet.title = "Sheet"
    # Add headers
    sheet['A1'] = 'اسم العميل'
    sheet['B1'] = 'النوع'
    sheet['C1'] = 'رقم الهاتف'
    # Add sample data
    sheet['A2'] = 'Test Customer'
    sheet['B2'] = 'عميل عادي'
    sheet['C2'] = '123456789'
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        test_workbook.save(tmp.name)
        test_file_path = tmp.name
    
    try:
        # Test normal operation
        success_count, error_count, errors = import_customers_from_excel(test_file_path)
        print(f"✓ import_customers_from_excel completed with {success_count} successes, {error_count} errors")
        
        # Test that we can now delete the file (meaning it's not locked)
        os.remove(test_file_path)
        print("✓ Successfully deleted test file after import operation (file was not locked)")
        return True
        
    except Exception as e:
        print(f"✗ Error during import test: {e}")
        # Try to manually delete the file for cleanup
        try:
            os.remove(test_file_path)
        except:
            pass
        return False


def main():
    """Run all tests"""
    print("Testing Excel import file handling fixes...")
    print("=" * 50)
    
    test1_result = test_restore_data_file_handling()
    test2_result = test_bulk_import_file_handling()
    
    print("\n" + "=" * 50)
    print("Test Results:")
    print(f"restore_data test: {'PASS' if test1_result else 'FAIL'}")
    print(f"bulk_import test: {'PASS' if test2_result else 'FAIL'}")
    
    if test1_result and test2_result:
        print("\n✓ All tests passed! The file locking issue should be resolved.")
        return True
    else:
        print("\n✗ Some tests failed. The file locking issue may still exist.")
        return False


if __name__ == "__main__":
    main()