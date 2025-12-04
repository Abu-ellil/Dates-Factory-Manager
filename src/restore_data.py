import openpyxl
from database import get_connection
from datetime import datetime
import os

def restore_from_excel(file_path, mode='merge'):
    """
    Restore data from Excel export file
    
    Args:
        file_path: Path to Excel file
        mode: 'merge' (add new data) or 'replace' (clear and restore)
    
    Returns: dict with restoration statistics
    """
    stats = {
        'customers': {'added': 0, 'skipped': 0, 'errors': []},
        'weighbridge': {'added': 0, 'skipped': 0, 'errors': []},
        'crates': {'added': 0, 'skipped': 0, 'errors': []},
        'finance': {'added': 0, 'skipped': 0, 'errors': []}
    }
    
    workbook = None
    try:
        # Load workbook in read-only and data-only mode to avoid locking issues
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        # Check if file has any valid sheets
        valid_sheets = ['العملاء (Customers)', 'الميزان (Weighbridge)', 'الصناديق (Crates)', 'المالية (Finance)']
        has_valid_sheets = any(sheet in workbook.sheetnames for sheet in valid_sheets)
        
        if not has_valid_sheets:
            return {'error': 'الملف لا يحتوي على أوراق عمل صالحة (العملاء، الميزان، الصناديق، المالية)'}
            
        conn = get_connection()
        cursor = conn.cursor()
        
        # If replace mode, clear existing data (except users)
        if mode == 'replace':
            cursor.execute('DELETE FROM finance')
            cursor.execute('DELETE FROM crates')
            cursor.execute('DELETE FROM weighbridge')
            cursor.execute('DELETE FROM customers')
            conn.commit()
        
        # Create customer name to ID mapping
        customer_map = {}
        
        # 1. Restore Customers
        if 'العملاء (Customers)' in workbook.sheetnames:
            sheet = workbook['العملاء (Customers)']
            # Process the sheet and immediately clear the reference
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or not row[1]:  # Skip empty rows
                        continue
                    
                    name = str(row[1]).strip()
                    customer_type = str(row[2]).strip() if row[2] else 'عميل عادي'
                    phone = str(row[3]).strip() if row[3] and row[3] != '-' else ''
                    
                    # Check if customer exists
                    existing = cursor.execute('SELECT id FROM customers WHERE name = ?', (name,)).fetchone()
                    
                    if existing:
                        customer_map[name] = existing['id']
                        stats['customers']['skipped'] += 1
                    else:
                        cursor.execute(
                            'INSERT INTO customers (name, type, phone) VALUES (?, ?, ?)',
                            (name, customer_type, phone)
                        )
                        customer_map[name] = cursor.lastrowid
                        stats['customers']['added'] += 1
                        
                except Exception as e:
                    stats['customers']['errors'].append(f"Row {row_idx}: {str(e)}")
            
            # Delete the sheet reference to avoid keeping workbook open
            del sheet
            conn.commit()
        
        # 2. Restore Weighbridge
        if 'الميزان (Weighbridge)' in workbook.sheetnames:
            sheet = workbook['الميزان (Weighbridge)']
            # Process the sheet and immediately clear the reference
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or not row[0]:
                        continue
                    
                    date = row[0]
                    if isinstance(date, datetime):
                        date = date.strftime('%Y-%m-%d')
                    else:
                        date = str(date)
                    
                    customer_name = str(row[1]).strip()
                    net_weight = float(row[2]) if row[2] else 0
                    price_per_qantar = float(row[3]) if row[3] else 0
                    total = float(row[4]) if row[4] else 0
                    
                    # Get customer ID
                    if customer_name in customer_map:
                        customer_id = customer_map[customer_name]
                    else:
                        customer = cursor.execute('SELECT id FROM customers WHERE name = ?', (customer_name,)).fetchone()
                        if customer:
                            customer_id = customer['id']
                            customer_map[customer_name] = customer_id
                        else:
                            stats['weighbridge']['errors'].append(f"Row {row_idx}: Customer '{customer_name}' not found")
                            continue
                    
                    cursor.execute('''
                        INSERT INTO weighbridge (date, customer_id, net_weight, price_per_qantar, total)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (date, customer_id, net_weight, price_per_qantar, total))
                    
                    stats['weighbridge']['added'] += 1
                    
                except Exception as e:
                    stats['weighbridge']['errors'].append(f"Row {row_idx}: {str(e)}")
            
            # Delete the sheet reference to avoid keeping workbook open
            del sheet
            conn.commit()
        
        # 3. Restore Crates
        if 'الصناديق (Crates)' in workbook.sheetnames:
            sheet = workbook['الصناديق (Crates)']
            # Process the sheet and immediately clear the reference
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or not row[0]:
                        continue
                    
                    date = row[0]
                    if isinstance(date, datetime):
                        date = date.strftime('%Y-%m-%d')
                    else:
                        date = str(date)
                    
                    customer_name = str(row[1]).strip()
                    crates_out = int(row[2]) if row[2] else 0
                    crates_returned = int(row[3]) if row[3] else 0
                    handler = str(row[4]).strip() if row[4] and row[4] != '-' else ''
                    notes = str(row[5]).strip() if row[5] and row[5] != '-' else ''
                    
                    # Get customer ID
                    if customer_name in customer_map:
                        customer_id = customer_map[customer_name]
                    else:
                        customer = cursor.execute('SELECT id FROM customers WHERE name = ?', (customer_name,)).fetchone()
                        if customer:
                            customer_id = customer['id']
                            customer_map[customer_name] = customer_id
                        else:
                            stats['crates']['errors'].append(f"Row {row_idx}: Customer '{customer_name}' not found")
                            continue
                    
                    cursor.execute('''
                        INSERT INTO crates (date, customer_id, crates_out, crates_returned, handler, notes)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (date, customer_id, crates_out, crates_returned, handler, notes))
                    
                    stats['crates']['added'] += 1
                    
                except Exception as e:
                    stats['crates']['errors'].append(f"Row {row_idx}: {str(e)}")
            
            # Delete the sheet reference to avoid keeping workbook open
            del sheet
            conn.commit()
        
        # 4. Restore Finance
        if 'المالية (Finance)' in workbook.sheetnames:
            sheet = workbook['المالية (Finance)']
            # Process the sheet and immediately clear the reference
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or not row[0]:
                        continue
                    
                    date = row[0]
                    if isinstance(date, datetime):
                        date = date.strftime('%Y-%m-%d')
                    else:
                        date = str(date)
                    
                    customer_name = str(row[1]).strip()
                    transaction_type = str(row[2]).strip() if row[2] else 'دفع'
                    amount_paid = float(row[3]) if row[3] else 0
                    amount_received = float(row[4]) if row[4] else 0
                    notes = str(row[5]).strip() if row[5] and row[5] != '-' else ''
                    
                    # Get customer ID
                    if customer_name in customer_map:
                        customer_id = customer_map[customer_name]
                    else:
                        customer = cursor.execute('SELECT id FROM customers WHERE name = ?', (customer_name,)).fetchone()
                        if customer:
                            customer_id = customer['id']
                            customer_map[customer_name] = customer_id
                        else:
                            stats['finance']['errors'].append(f"Row {row_idx}: Customer '{customer_name}' not found")
                            continue
                    
                    cursor.execute('''
                        INSERT INTO finance (date, customer_id, transaction_type, amount_paid, amount_received, notes)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (date, customer_id, transaction_type, amount_paid, amount_received, notes))
                    
                    stats['finance']['added'] += 1
                    
                except Exception as e:
                    stats['finance']['errors'].append(f"Row {row_idx}: {str(e)}")
            
            # Delete the sheet reference to avoid keeping workbook open
            del sheet
            conn.commit()
        
        conn.close()
        
        return stats
        
    except Exception as e:
        return {'error': str(e)}
    finally:
        # Ensure workbook is closed in all cases
        if workbook is not None:
            try:
                workbook.close()
            except:
                pass  # Already closed or other issue

if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        mode = sys.argv[2] if len(sys.argv) > 2 else 'merge'
        
        print(f"🔄 Restoring data from: {file_path}")
        print(f"📋 Mode: {mode}")
        print()
        
        stats = restore_from_excel(file_path, mode)
        
        if 'error' in stats:
            print(f"❌ Error: {stats['error']}")
        else:
            print("✅ Restoration completed!")
            print()
            print(f"📊 Customers: {stats['customers']['added']} added, {stats['customers']['skipped']} skipped")
            print(f"⚖️  Weighbridge: {stats['weighbridge']['added']} added")
            print(f"📦 Crates: {stats['crates']['added']} added")
            print(f"💰 Finance: {stats['finance']['added']} added")
            
            # Show errors if any
            for category in ['customers', 'weighbridge', 'crates', 'finance']:
                if stats[category]['errors']:
                    print(f"\n⚠️  {category.capitalize()} errors:")
                    for error in stats[category]['errors'][:5]:
                        print(f"   - {error}")
    else:
        print("Usage: python restore_data.py <excel_file_path> [mode]")
        print("Modes: merge (default) or replace")
