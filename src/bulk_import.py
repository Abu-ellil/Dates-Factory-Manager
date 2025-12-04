import openpyxl
from database import get_connection

def import_customers_from_excel(file_path):
    """
    Import customers from Excel file
    Expected columns: اسم العميل, النوع, رقم الهاتف
    Or: Name, Type, Phone
    
    Returns: (success_count, error_count, errors_list)
    """
    workbook = None
    try:
        # Load workbook in read-only and data-only mode to avoid locking issues
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        
        conn = get_connection()
        success_count = 0
        error_count = 0
        errors = []
        
        # Try to find header row
        headers = []
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            if row[0] and ('اسم' in str(row[0]).lower() or 'name' in str(row[0]).lower()):
                headers = [str(cell).strip() if cell else '' for cell in row]
                break
        
        if not headers:
            # Assume first row is data, not headers
            headers = ['Name', 'Type', 'Phone']
            start_row = 1
        else:
            start_row = sheet.min_row + headers.index(headers[0]) + 1
        
        # Find column indices
        name_col = None
        type_col = None
        phone_col = None
        
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if 'اسم' in header_lower or 'name' in header_lower:
                name_col = idx
            elif 'نوع' in header_lower or 'type' in header_lower:
                type_col = idx
            elif 'هاتف' in header_lower or 'phone' in header_lower or 'رقم' in header_lower:
                phone_col = idx
        
        # If columns not found, assume order: Name, Type, Phone
        if name_col is None:
            name_col = 0
        if type_col is None:
            type_col = 1
        if phone_col is None:
            phone_col = 2
        
        # Import data
        for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            try:
                if not row or not row[name_col]:
                    continue
                
                name = str(row[name_col]).strip()
                customer_type = str(row[type_col]).strip() if type_col < len(row) and row[type_col] else 'عميل عادي'
                phone = str(row[phone_col]).strip() if phone_col < len(row) and row[phone_col] else ''
                
                # Validate
                if not name:
                    errors.append(f"Row {row_idx}: اسم العميل فارغ")
                    error_count += 1
                    continue
                
                # Insert into database
                conn.execute(
                    'INSERT OR IGNORE INTO customers (name, type, phone) VALUES (?, ?, ?)',
                    (name, customer_type, phone)
                )
                
                if conn.total_changes > 0:
                    success_count += 1
                else:
                    errors.append(f"Row {row_idx}: العميل '{name}' موجود بالفعل")
                    error_count += 1
                    
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                error_count += 1
        
        # Delete the sheet reference to avoid keeping workbook open
        del sheet
        
        conn.commit()
        conn.close()
        
        return success_count, error_count, errors
        
    except Exception as e:
        return 0, 0, [f"Error reading file: {str(e)}"]
    finally:
        # Ensure workbook is closed in all cases
        if workbook is not None:
            try:
                workbook.close()
            except:
                pass  # Already closed or other issue
            # Explicitly set to None to remove reference
            workbook = None

if __name__ == '__main__':
    # Test import
    import sys
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        success, errors, error_list = import_customers_from_excel(file_path)
        print(f"✅ Imported: {success}")
        print(f"❌ Errors: {errors}")
        if error_list:
            print("\nError details:")
            for error in error_list[:10]:  # Show first 10 errors
                print(f"  - {error}")
    else:
        print("Usage: python bulk_import.py <excel_file_path>")
